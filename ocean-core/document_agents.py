"""
DOCUMENT AGENTS - Industrial-grade document generation for Clisonix Ocean.
Supports PDF, Excel, CSV, and Report formats via contract-governed pipeline.
"""

import csv
import io
import json
import logging
from datetime import datetime
from enum import Enum
from typing import Any, Dict, List, Optional, Union

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


class DocumentFormat(Enum):
    """Supported document output formats."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    REPORT = "report"
    JSON = "json"


class DocumentAgent:
    """Base agent for document generation."""
    
    def __init__(self, format_type: DocumentFormat):
        self.format_type = format_type
        self.format_name = format_type.value
        
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate document based on contract and query."""
        raise NotImplementedError


class ExcelAgent(DocumentAgent):
    """Excel/CSV document generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.EXCEL)
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate Excel or CSV document."""
        if not HAS_OPENPYXL:
            return {
                "success": False,
                "errors": ["openpyxl not installed"],
                "validation_status": "failed"
            }
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # Header
            ws['A1'] = "Clisonix Cloud - Document Report"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Generated: {datetime.now().isoformat()}"
            ws['A3'] = f"Query: {query}"
            ws['A3'].font = Font(italic=True)
            
            # Content from contract
            row = 5
            if hasattr(contract, 'get_data'):
                data = contract.get_data()
                for key, value in data.items():
                    ws[f'A{row}'] = key
                    ws[f'B{row}'] = str(value)
                    row += 1
            
            # Auto-width columns
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 50
            
            # Generate bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return {
                "success": True,
                "validation_status": "passed",
                "errors": [],
                "document": {
                    "type": "workbook",
                    "format": "xlsx",
                    "data": output.getvalue().hex(),
                    "rows": row - 5,
                    "columns": 2
                },
                "provenance": {
                    "agent": "ExcelAgent",
                    "timestamp": datetime.utcnow().isoformat()
                }
            }
        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class PDFAgent(DocumentAgent):
    """PDF document generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.PDF)
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate PDF document."""
        if not HAS_REPORTLAB:
            return {
                "success": False,
                "errors": ["reportlab not installed"],
                "validation_status": "failed"
            }
        
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            story = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor("#1F2937"),
                spaceAfter=30,
                alignment=1
            )
            
            # Title
            story.append(Paragraph("Clisonix Cloud Report", title_style))
            story.append(Spacer(1, 0.5*inch))
            
            # Metadata
            story.append(Paragraph(f"<b>Query:</b> {query}", styles['Normal']))
            story.append(Paragraph(f"<b>Generated:</b> {datetime.now().isoformat()}", styles['Normal']))
            story.append(Paragraph(f"<b>Language:</b> {language}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Content from contract
            if hasattr(contract, 'get_summary'):
                summary = contract.get_summary()
                story.append(Paragraph("<b>Summary:</b>", styles['Heading2']))
                story.append(Paragraph(summary, styles['Normal']))
            
            doc.build(story)
            output.seek(0)
            
            return {
                "success": True,
                "validation_status": "passed",
                "errors": [],
                "document": {
                    "type": "pdf",
                    "format": "pdf",
                    "data": output.getvalue().hex(),
                    "size_bytes": len(output.getvalue())
                },
                "provenance": {
                    "agent": "PDFAgent",
                    "timestamp": datetime.utcnow().isoformat()
                }
            }
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class ReportAgent(DocumentAgent):
    """Structured report generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.REPORT)
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate structured report."""
        try:
            report = {
                "title": "Clisonix Cloud - Industrial Report",
                "query": query,
                "language": language,
                "generated_at": datetime.utcnow().isoformat(),
                "contract_type": getattr(contract, 'contract_type', 'generic'),
                "sections": []
            }
            
            # Extract sections from contract
            if hasattr(contract, 'get_sections'):
                report["sections"] = contract.get_sections()
            else:
                report["sections"] = [
                    {
                        "title": "Executive Summary",
                        "content": f"Report for query: {query}"
                    },
                    {
                        "title": "Details",
                        "content": "This is a generated report from Clisonix Ocean."
                    }
                ]
            
            return {
                "success": True,
                "validation_status": "passed",
                "errors": [],
                "document": report,
                "provenance": {
                    "agent": "ReportAgent",
                    "timestamp": datetime.utcnow().isoformat()
                }
            }
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class CSVAgent(DocumentAgent):
    """CSV document generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.CSV)
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate CSV document."""
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            writer.writerow(["Clisonix Cloud - Data Export"])
            writer.writerow([f"Generated: {datetime.utcnow().isoformat()}"])
            writer.writerow([f"Query: {query}"])
            writer.writerow([f"Language: {language}"])
            writer.writerow([])
            
            # Data rows from contract
            if hasattr(contract, 'get_data'):
                data = contract.get_data()
                writer.writerow(["Key", "Value"])
                for key, value in data.items():
                    writer.writerow([key, value])
            else:
                writer.writerow(["No data"])
            
            csv_content = output.getvalue()
            
            return {
                "success": True,
                "validation_status": "passed",
                "errors": [],
                "document": {
                    "type": "csv",
                    "format": "csv",
                    "data": csv_content,
                    "rows": len(csv_content.split('\n'))
                },
                "provenance": {
                    "agent": "CSVAgent",
                    "timestamp": datetime.utcnow().isoformat()
                }
            }
        except Exception as e:
            logger.error(f"CSV generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class VideoAgent(DocumentAgent):
    """Video document generation agent (Blerina + Animated)."""
    
    def __init__(self):
        super().__init__(DocumentFormat.REPORT)
        self.format_name = "video"
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate video from contract."""
        try:
            # Import video generators
            try:
                from video_generator_blerina import ScriptGenerator as BberinaScriptGen
                from video_generator_blerina import VideoProject, VideoStyle
                HAS_BLERINA = True
            except ImportError:
                HAS_BLERINA = False
                logger.warning("Blerina video generator not available")
            
            if not HAS_BLERINA:
                return {
                    "success": False,
                    "errors": ["Blerina video generator not available"],
                    "validation_status": "unavailable"
                }
            
            # Generate video script and project metadata
            video_project = {
                "title": getattr(contract, 'title', 'Generated Video'),
                "topic": query,
                "style": "educational",
                "language": language,
                "sections": [],
                "metadata": {
                    "source": "ocean_document_generation",
                    "contract_type": getattr(contract, 'contract_type', 'generic'),
                    "generated_at": datetime.utcnow().isoformat()
                }
            }
            
            # Extract content from contract
            if hasattr(contract, 'get_sections'):
                sections = contract.get_sections()
                for i, section in enumerate(sections):
                    video_project["sections"].append({
                        "index": i,
                        "title": section.get("title", "Section"),
                        "narration": section.get("content", ""),
                        "duration_estimate": len(section.get("content", "").split()) / 130  # ~130 words/min
                    })
            
            return {
                "success": True,
                "validation_status": "ready_for_generation",
                "errors": [],
                "document": {
                    "type": "video_project",
                    "format": "mp4",
                    "project": video_project,
                    "generators": ["blerina", "animated"],
                    "estimated_duration_seconds": sum(s.get("duration_estimate", 5) for s in video_project["sections"])
                },
                "provenance": {
                    "agent": "VideoAgent",
                    "timestamp": datetime.utcnow().isoformat(),
                    "backend": ["video_generator_blerina", "video_generator_animated"]
                }
            }
        except Exception as e:
            logger.error(f"Video generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class VoiceAgent(DocumentAgent):
    """Voice/Audio document generation agent (TTS + Voice synthesis)."""
    
    def __init__(self):
        super().__init__(DocumentFormat.REPORT)
        self.format_name = "voice"
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate voice/audio from contract."""
        try:
            voice_project = {
                "title": getattr(contract, 'title', 'Generated Voice Document'),
                "query": query,
                "language": language,
                "voice_styles": ["professional", "friendly", "narrator"],
                "segments": [],
                "metadata": {
                    "tts_backend": "coqui_tts",
                    "sample_rate": 22050,
                    "format": "wav",
                    "generated_at": datetime.utcnow().isoformat()
                }
            }
            
            # Extract narration from contract
            if hasattr(contract, 'get_summary'):
                summary = contract.get_summary()
                voice_project["segments"].append({
                    "type": "summary",
                    "text": summary,
                    "voice_style": "professional",
                    "duration_estimate": len(summary.split()) / 130  # ~130 words/min
                })
            
            if hasattr(contract, 'get_sections'):
                sections = contract.get_sections()
                for i, section in enumerate(sections):
                    voice_project["segments"].append({
                        "type": "section",
                        "index": i,
                        "title": section.get("title", "Section"),
                        "text": section.get("content", ""),
                        "voice_style": "friendly" if i % 2 == 0 else "narrator",
                        "duration_estimate": len(section.get("content", "").split()) / 130
                    })
            
            total_duration = sum(s.get("duration_estimate", 0) for s in voice_project["segments"])
            
            return {
                "success": True,
                "validation_status": "ready_for_generation",
                "errors": [],
                "document": {
                    "type": "voice_project",
                    "format": "wav",
                    "project": voice_project,
                    "total_duration_seconds": total_duration,
                    "segments_count": len(voice_project["segments"])
                },
                "provenance": {
                    "agent": "VoiceAgent",
                    "timestamp": datetime.utcnow().isoformat(),
                    "tts_engine": "coqui_tts",
                    "backends": "ocean_nanogrid /api/ocean/tts"
                }
            }
        except Exception as e:
            logger.error(f"Voice generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class MusicAgent(DocumentAgent):
    """Music/soundtrack generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.REPORT)
        self.format_name = "music"
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate music composition from contract."""
        try:
            # Check for music21
            try:
                import music21
                HAS_MUSIC21 = True
            except ImportError:
                HAS_MUSIC21 = False
                logger.warning("music21 not available")
            
            music_project = {
                "title": getattr(contract, 'title', 'Generated Music'),
                "query": query,
                "genre": getattr(contract, 'music_genre', 'ambient'),
                "bpm": getattr(contract, 'bpm', 120),
                "key": getattr(contract, 'key', 'C'),
                "time_signature": getattr(contract, 'time_signature', '4/4'),
                "duration_seconds": getattr(contract, 'duration_seconds', 300),
                "instruments": getattr(contract, 'instruments', ['piano']),
                "mood": getattr(contract, 'mood', 'calm'),
                "language": language,
                "metadata": {
                    "music_engine": "music21" if HAS_MUSIC21 else "procedural",
                    "format": "midi",
                    "generated_at": datetime.utcnow().isoformat()
                }
            }
            
            return {
                "success": True,
                "validation_status": "ready_for_generation",
                "errors": [],
                "document": {
                    "type": "music_project",
                    "format": "midi",
                    "project": music_project,
                    "total_duration_seconds": music_project["duration_seconds"]
                },
                "provenance": {
                    "agent": "MusicAgent",
                    "timestamp": datetime.utcnow().isoformat(),
                    "music_engine": "music21" if HAS_MUSIC21 else "procedural"
                }
            }
        except Exception as e:
            logger.error(f"Music generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class PaintingAgent(DocumentAgent):
    """Image/painting generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.REPORT)
        self.format_name = "painting"
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate painting/image from contract."""
        try:
            # Check for PIL
            try:
                from PIL import Image
                HAS_PIL = True
            except ImportError:
                HAS_PIL = False
                logger.warning("PIL not available")
            
            painting_project = {
                "title": getattr(contract, 'title', 'Generated Painting'),
                "query": query,
                "style": getattr(contract, 'style', 'digital_art'),
                "theme": getattr(contract, 'theme', 'abstract'),
                "width": getattr(contract, 'width', 1920),
                "height": getattr(contract, 'height', 1080),
                "color_palette": getattr(contract, 'color_palette', ['#FF6B6B', '#4ECDC4']),
                "lighting": getattr(contract, 'lighting', 'natural'),
                "composition": getattr(contract, 'composition', 'balanced'),
                "language": language,
                "metadata": {
                    "renderer": "PIL" if HAS_PIL else "procedural",
                    "format": "png",
                    "generated_at": datetime.utcnow().isoformat()
                }
            }
            
            return {
                "success": True,
                "validation_status": "ready_for_generation",
                "errors": [],
                "document": {
                    "type": "painting_project",
                    "format": "png",
                    "project": painting_project,
                    "resolution": f"{painting_project['width']}x{painting_project['height']}"
                },
                "provenance": {
                    "agent": "PaintingAgent",
                    "timestamp": datetime.utcnow().isoformat(),
                    "renderer": "PIL" if HAS_PIL else "procedural"
                }
            }
        except Exception as e:
            logger.error(f"Painting generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


class AnimationAgent(DocumentAgent):
    """Animation/motion graphics generation agent."""
    
    def __init__(self):
        super().__init__(DocumentFormat.REPORT)
        self.format_name = "animation"
    
    def generate_document(self, contract: Any, query: str, language: str = "en") -> Dict[str, Any]:
        """Generate animation from contract."""
        try:
            # Check for OpenCV
            try:
                import cv2
                HAS_OPENCV = True
            except ImportError:
                HAS_OPENCV = False
                logger.warning("OpenCV not available")
            
            animation_project = {
                "title": getattr(contract, 'title', 'Generated Animation'),
                "query": query,
                "style": getattr(contract, 'animation_style', 'modern'),
                "fps": getattr(contract, 'fps', 30),
                "duration_seconds": getattr(contract, 'duration_seconds', 60),
                "resolution": getattr(contract, 'resolution', '1080p'),
                "frames_count": len(getattr(contract, 'frames', [])),
                "language": language,
                "metadata": {
                    "renderer": "OpenCV" if HAS_OPENCV else "procedural",
                    "format": "mp4",
                    "generated_at": datetime.utcnow().isoformat()
                }
            }
            
            return {
                "success": True,
                "validation_status": "ready_for_generation",
                "errors": [],
                "document": {
                    "type": "animation_project",
                    "format": "mp4",
                    "project": animation_project,
                    "total_frames": animation_project["frames_count"] or animation_project["fps"] * animation_project["duration_seconds"]
                },
                "provenance": {
                    "agent": "AnimationAgent",
                    "timestamp": datetime.utcnow().isoformat(),
                    "renderer": "OpenCV" if HAS_OPENCV else "procedural"
                }
            }
        except Exception as e:
            logger.error(f"Animation generation failed: {e}")
            return {
                "success": False,
                "errors": [str(e)],
                "validation_status": "failed"
            }


# Global agent registry
_AGENTS = {
    "pdf": PDFAgent(),
    "excel": ExcelAgent(),
    "csv": CSVAgent(),
    "report": ReportAgent(),
    "video": VideoAgent(),
    "voice": VoiceAgent(),
    "music": MusicAgent(),
    "painting": PaintingAgent(),
    "animation": AnimationAgent(),
}


def get_agent(format_name: str) -> Optional[DocumentAgent]:
    """Get agent by format name."""
    return _AGENTS.get(format_name.lower())


def list_agents() -> List[Dict[str, Any]]:
    """List available document agents."""
    return [
        {
            "name": "pdf",
            "format": "PDF",
            "description": "Portable Document Format",
            "available": HAS_REPORTLAB
        },
        {
            "name": "excel",
            "format": "XLSX/CSV",
            "description": "Microsoft Excel / Comma-Separated Values",
            "available": HAS_OPENPYXL
        },
        {
            "name": "report",
            "format": "JSON Report",
            "description": "Structured report format",
            "available": True
        },
        {
            "name": "csv",
            "format": "CSV",
            "description": "Comma-Separated Values",
            "available": True
        },
        {
            "name": "video",
            "format": "MP4 Video",
            "description": "Video generation via Blerina & Animated generators",
            "available": True,
            "backends": ["video_generator_blerina", "video_generator_animated"],
            "features": ["blerina_script", "animated_motion_graphics", "tts_narration", "unlimited_concepts"]
        },
        {
            "name": "voice",
            "format": "WAV Audio",
            "description": "Voice/TTS generation via Coqui TTS",
            "available": True,
            "backends": ["coqui_tts", "ocean_nanogrid_tts"],
            "features": ["multi_voice_styles", "multilingual", "fast_generation"]
        },
        {
            "name": "music",
            "format": "MIDI/MP3 Music",
            "description": "Music composition generation via music21",
            "available": True,
            "backends": ["music21", "procedural"],
            "features": ["multiple_genres", "bpm_control", "instrument_selection", "mood_based"]
        },
        {
            "name": "painting",
            "format": "PNG/JPG Image",
            "description": "Image/painting generation via PIL and procedural methods",
            "available": True,
            "backends": ["PIL", "procedural"],
            "features": ["style_selection", "color_palettes", "composition_control", "theme_based"]
        },
        {
            "name": "animation",
            "format": "MP4 Animation",
            "description": "Animation/motion graphics generation via OpenCV",
            "available": True,
            "backends": ["OpenCV", "procedural"],
            "features": ["keyframe_animation", "transitions", "effects", "fps_control"]
        }
    ]
